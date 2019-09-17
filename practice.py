import openpyxl
import xlrd
import csv
import os

def receipts_view(request):
    if request.method == "POST":
        excel_files = request.FILES['excel_files']
        print(excel_files)
        wb = openpyxl.load_workbook(excel_files)
        worksheet = wb.active
        print(worksheet)
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        return render(request, 'main/receipts_view.html',{'excel_data':excel_data})

    else:
        return render(request,'main/receipts_view.html')


def receipts(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            # if 'overview' in request.POST:
            #     if Upload_Excel.objects.all() == []:
            #         return render(request,'main/receipts.html',{'error_message_1':"없음"})
            #     else:
            #         print(Upload_Excel.objects.all())
            #         return render(request,'main.receipts.html')

            if 'initial' in request.POST:
                if Upload_Excel.objects.all() == []:
                    upload = form.save(commit=False)
                    upload.author = request.user
                    upload.published_date = timezone.now()
                    upload.title = request.user
                    upload.save()


                    excel_file = upload.file
                    excel_data = excel_read(excel_file)

                    return render(request, 'main/receipts.html', {'excel_data': excel_data})

                else:
                    Upload_Excel.objects.all().delete()
                    upload = form.save(commit=False)
                    upload.author = request.user
                    upload.published_date = timezone.now()
                    upload.title = request.user
                    upload.save()

                    excel_file = upload.file
                    excel_data = excel_read(excel_file)

                    return render(request, 'main/receipts.html', {'excel_data':excel_data})

            # elif 'delete' in request.POST:
            #     if os.path.exists(settings.MEDIA_ROOT):
            #         for file in os.scandir(settings.MEDIA_ROOT):
            #             os.remove(file.path)
            #         return render(request, 'main/receipts.html', {'confirm_message': "삭제 완료"})
            #     else:
            #         return redirect(receipts)


            elif 'add' in request.POST:
                if Upload_Excel.objects.all() == []:
                    return render(request,'main/receipts.html',{'error_message':"초기 업로드를 먼저 진행하세요."})
                else:
                    excel_file1 = Upload_Excel.objects.all().get().file
                    excel_data1 = excel_read(excel_file1)

                    upload = form.save(commit=False)
                    upload.author = request.user
                    upload.published_date = timezone.now()
                    upload.title = request.user
                    # upload.save()

                    excel_file2 = upload.file
                    excel_data2 = excel_read(excel_file2)
                    excel_add(excel_data1, excel_data2, upload.title, upload.root)
                    excel_data = excel_read(settings.MEDIA_ROOT + '\\' + str(upload.root) + '\\' + str(upload.title) + '.xlsx')
                    upload.file = settings.MEDIA_ROOT + '\\' + str(upload.root) + '\\' + str(upload.title) + '.xlsx'
                    upload.save()
                    objects = Upload_Excel.objects.all()
                    objects[0].delete()
                    return render(request,'main/receipts.html',{'excel_data':excel_data})

        else:
            return redirect('receipts')

    else:
        form = UploadForm()

    return render(request,'main/receipts.html',{'form':form})



def excel_read(excel_file):
    excel_file = 'D:\\Project\\MD(Rohm_International)\\Room\\7\\prac.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb.active
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return excel_data

def excel_add(excel_file1, excel_file2, title, root):
    excel_data1 = excel_file1
    excel_data2 = excel_file2
    excel_data = excel_data1+excel_data2

    wb_new = openpyxl.Workbook()
    ws_write = wb_new.active
    for r in range(len(excel_data)):
        for c in range(len(excel_data[0])):
            ws_write.cell(row=r+1,column=c+1).value = excel_data[r][c]
    wb_new.save(filename=settings.MEDIA_ROOT + '\\' + str(root) + '\\' + str(title) + '.xlsx')
    print('완료')


def main():
    excel_file = 'D:\\Project\\MD(Rohm_International)\\Room\\7\\prac.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    worksheet = wb.active
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return excel_data

    # excel_file1 = 'D:\\Project\\MD(Rohm_International)\\Room\\7\\prac.xlsx'
    # wb = openpyxl.load_workbook(excel_file1)
    # worksheet = wb.active
    # excel_data1 = list()
    # for row in worksheet.iter_rows():
    #     row_data = list()
    #     for cell in row:
    #         row_data.append(str(cell.value))
    #     excel_data1.append(row_data)
    #
    # excel_file2 = 'D:\\Project\\MD(Rohm_International)\\Room\\7\\prac2.xlsx'
    # wb = openpyxl.load_workbook(excel_file2)
    # worksheet = wb.active
    # excel_data2 = list()
    # for row in worksheet.iter_rows():
    #     row_data = list()
    #     for cell in row:
    #         row_data.append(str(cell.value))
    #     excel_data2.append(row_data)
    #
    # excel_data = excel_data1 + excel_data2
    #
    # wb_new = openpyxl.Workbook()
    # ws_write = wb_new.active
    # for r in range(len(excel_data)):
    #     for c in range(len(excel_data[0])):
    #         ws_write.cell(row=r+1,column=c+1).value = excel_data[r][c]
    # wb_new.save(filename='D:\\Project\\MD(Rohm_International)\\Room\\7\\11.xlsx')

    # excel_file = 'D:\\Project\\MD(Rohm_International)\\Room\\1\\1_1.csv'
    #
    # fname, ext = os.path.splitext(excel_file)
    #
    # if ext=='.xls':
    #     wb = xlrd.open_workbook(excel_file)
    #     worksheet = wb.sheet_by_index(0)
    #     nrows = worksheet.nrows
    #     row_val = []
    #     for row_num in range(nrows):
    #         row_val.append(worksheet.row_values(row_num))
    #
    # elif ext=='.xlsx':
    #     wb = openpyxl.load_workbook(excel_file)
    #     worksheet = wb.active
    #     print(worksheet)
    #     excel_data = list()
    #     for row in worksheet.iter_rows():
    #         row_data = list()
    #         for cell in row:
    #             row_data.append(str(cell.value))
    #         excel_data.append(row_data)
    #
    # elif ext=='.csv':
    #     f = open(excel_file,'r',encoding='utf-8-sig')
    #     wb = csv.reader(f)
    #     val = []
    #     for line in wb:
    #         val.append(line)
    #
    # excel_file1 = 'D:\\Project\\MD(Rohm_International)\\Room\\11\\2_3_1.xlsx'
    # excel_file2 = 'D:\\Project\\MD(Rohm_International)\\Room\\1\\3_1.xls'
    #
    # a = openpyxl.load_workbook(excel_file1)
    # b = xlrd.open_workbook(excel_file2)
    #
    #
    #
    # if excel_file.name.endswith('.csv'):
    #     wb = xlrd.open_workbook(excel_file)
    #     worksheet = wb.sheet_by_index(0)
    #     nrows = worksheet.nrows
    #     row_val = []
    #     for row_num in range(nrows):
    #         row_val.append(worksheet.row_values(row_num))
    #
    # elif excel_file.name.endswith('.xlsx'):
    #     wb = openpyxl.load_workbook(excel_file)
    #     worksheet = wb.active
    #     print(worksheet)
    #     excel_data = list()
    #     for row in worksheet.iter_rows():
    #         row_data = list()
    #         for cell in row:
    #             row_data.append(str(cell.value))
    #         excel_data.append(row_data)
    #
    # elif excel_file.name.endswith('.csv'):
    #     f = open(excel_file,'r',encoding='utf-8-sig')
    #     wb = csv.reader(f)
    #     val = []
    #     for line in wb:
    #         val.append(line)
    #
    # excel_file = 'D:\\Project\\MD(Rohm_International)\\Room\\11\\2_3_1.xlsx'
    #
    # fname, ext = os.path.splitext(excel_file)
    #
    # if ext=='.xls':
    #     wb = xlrd.open_workbook(excel_file)
    #     worksheet = wb.sheet_by_index(0)
    #     nrows = worksheet.nrows
    #     row_val = []
    #     for row_num in range(nrows):
    #         str(row_val.append(worksheet.row_values(row_num)))
    #
    # elif ext=='.xlsx':
    #     wb = openpyxl.load_workbook(excel_file)
    #     worksheet = wb.active
    #     print(worksheet)
    #     excel_data = list()
    #     for row in worksheet.iter_rows():
    #         row_data = list()
    #         for cell in row:
    #             row_data.append(str(cell.value))
    #         excel_data.append(row_data)
    #
    # elif ext=='.csv':
    #     # f = open(excel_file,'r',encoding='utf-8-sig')
    #     wb = csv.reader(excel_file)
    #     val = []
    #     for line in wb:
    #         val.append(str(line))
    #
    # excel_file = 'D:\\Project\\MD(Rohm_International)\\Room\\7\\prac.xlsx'
    # wb = openpyxl.load_workbook(excel_file)
    # worksheet = wb.active
    # print(worksheet)
    # excel_data = list()
    # for row in worksheet.iter_rows():
    #     row_data = list()
    #     for cell in row:
    #         row_data.append(str(cell.value))
    #     excel_data.append(row_data)


    print('g')

if __name__=='__main__':
    main()

